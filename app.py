import re
import pathlib
import streamlit as st
import pandas as pd
import io
import plotly.express as px
from datetime import datetime, date, time, timedelta
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, Reference

st.set_page_config(
    page_title="Agent AUX Analyzer",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ──────────────────────────────────────────────────────────────────────────────
# Styling
# ──────────────────────────────────────────────────────────────────────────────
st.markdown(
    """
    <style>
        .block-container { padding-top: 4rem; }
        .main-title {
            font-size: 2rem;
            font-weight: 700;
            margin-bottom: 0.2rem;
        }
        .sub-title {
            font-size: 0.95rem;
            opacity: 0.65;
            margin-bottom: 1.5rem;
        }
        .section-header {
            font-size: 1.2rem;
            font-weight: 600;
            border-bottom: 2px solid rgba(128,128,128,0.35);
            padding-bottom: 4px;
            margin-top: 1rem;
        }
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown('<div class="main-title">📊 CS Agent AUX Code Analyzer</div>', unsafe_allow_html=True)
st.markdown(
    '<div class="sub-title">Dynamically explore which AUX codes any agent was on during any time slot, '
    "with exact overlap durations and visual breakdowns. "
    "<strong>All dates &amp; times are displayed in IST (UTC+5:30).</strong></div>",
    unsafe_allow_html=True,
)

# ──────────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────────
def fmt_duration(seconds: float) -> str:
    """Format seconds to HH:MM:SS string."""
    seconds = int(max(0, seconds))
    h, rem = divmod(seconds, 3600)
    m, s = divmod(rem, 60)
    return f"{h:02d}:{m:02d}:{s:02d}"


def _merge_intervals(intervals: list) -> list:
    """Merge overlapping (start, end) datetime pairs. Returns sorted, merged list."""
    if not intervals:
        return []
    ivs = sorted(intervals, key=lambda x: x[0])
    merged = [list(ivs[0])]
    for s, e in ivs[1:]:
        if s < merged[-1][1]:               # overlaps → extend
            merged[-1][1] = max(merged[-1][1], e)
        else:
            merged.append([s, e])
    return [(s, e) for s, e in merged]


def _union_secs(intervals: list) -> float:
    """Total seconds covered by the union of (start, end) datetime pairs."""
    return sum((e - s).total_seconds() for s, e in _merge_intervals(intervals))


def _subtract_intervals(ivs_a: list, ivs_b: list) -> list:
    """Subtract ivs_b from ivs_a. Returns remaining (start, end) pairs."""
    if not ivs_b or not ivs_a:
        return list(ivs_a)
    remaining = list(ivs_a)
    for b_s, b_e in _merge_intervals(ivs_b):
        new = []
        for a_s, a_e in remaining:
            if b_e <= a_s or b_s >= a_e:
                new.append((a_s, a_e))
            else:
                if a_s < b_s:
                    new.append((a_s, b_s))
                if b_e < a_e:
                    new.append((b_e, a_e))
        remaining = new
    return remaining


_SESSION_AUX = frozenset({"in call - working", "in call- working"})
_SP_AUX      = "special project"


def _allocate_aux_secs(triples: list) -> dict:
    """Allocate slot time to AUX labels without double-counting overlapping intervals.

    Priority: specific AUX codes  >  In Call - Working (session)  >  Special Project
    Special Project is included only when fully isolated (no overlap with any other record).

    Args:
        triples: list of (label_str, start_datetime, end_datetime)
    Returns:
        {label: seconds} — only labels with seconds > 0 are included.
    """
    specific, sessions, special = [], [], []
    for lbl, s, e in triples:
        low = str(lbl).lower().strip()
        if low == _SP_AUX:
            special.append((lbl, s, e))
        elif low in _SESSION_AUX:
            sessions.append((lbl, s, e))
        else:
            specific.append((lbl, s, e))

    result  = {}
    claimed = []            # merged list of already-claimed (start, end) intervals

    def _claim_group(group):
        by_label = {}
        for lbl, s, e in group:
            by_label.setdefault(lbl, []).append((s, e))
        for label, ivs in by_label.items():
            unique = _subtract_intervals(_merge_intervals(ivs), claimed)
            secs   = sum((e - s).total_seconds() for s, e in unique)
            if secs > 0:
                result[label] = result.get(label, 0) + secs
            if unique:
                claimed.extend(unique)
                claimed[:] = _merge_intervals(claimed)

    _claim_group(specific)
    _claim_group(sessions)

    if special:
        sp_by_label = {}
        for lbl, s, e in special:
            sp_by_label.setdefault(lbl, []).append((s, e))
        for label, sp_ivs in sp_by_label.items():
            # Subtract all already-claimed time from Special Project;
            # only the uncovered residual is attributed to Special Project.
            unique = _subtract_intervals(_merge_intervals(sp_ivs), claimed)
            secs   = sum((e - s).total_seconds() for s, e in unique)
            if secs > 0:
                result[label] = result.get(label, 0) + secs
                if unique:
                    claimed.extend(unique)
                    claimed[:] = _merge_intervals(claimed)

    return result


def _parse_dur_secs(val) -> float | None:
    """Parse a Duration cell value to seconds.
    Handles datetime.time objects, timedelta, and HH:MM:SS / H:MM:SS strings."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, timedelta):
        return val.total_seconds()
    if hasattr(val, "hour"):          # datetime.time
        return val.hour * 3600 + val.minute * 60 + val.second
    try:
        parts = str(val).strip().split(":")
        if len(parts) == 3:
            return int(parts[0]) * 3600 + int(parts[1]) * 60 + float(parts[2])
        if len(parts) == 2:
            return int(parts[0]) * 60 + float(parts[1])
    except Exception:
        pass
    return None


def find_col(df: pd.DataFrame, candidates: list[str]) -> str | None:
    """Case-insensitive column lookup from a list of candidate names."""
    lookup = {c.lower().strip().replace(" ", ""): c for c in df.columns}
    for cand in candidates:
        key = cand.lower().replace(" ", "")
        if key in lookup:
            return lookup[key]
    return None


def parse_time_slot(slot_str: str):
    """Parse slot strings like '6 PM-7PM', '10 PM to 11 PM', '8PM to 9 PM'.
    Returns (start_time, end_time) or None if unparseable."""
    s = str(slot_str).strip()
    parts = re.split(r'\s*(?:\bto\b|-)\s*', s, maxsplit=1, flags=re.IGNORECASE)
    if len(parts) != 2:
        return None

    def _t(ts: str):
        ts = ts.strip()
        m = re.match(r'(\d{1,2})(?::(\d{2}))?\s*(AM|PM)', ts, re.IGNORECASE)
        if not m:
            return None
        h, mi, period = int(m.group(1)), int(m.group(2) or 0), m.group(3).upper()
        if period == "PM" and h != 12:
            h += 12
        if period == "AM" and h == 12:
            h = 0
        return time(h, mi)

    s_time, e_time = _t(parts[0]), _t(parts[1])
    return (s_time, e_time) if s_time and e_time else None


# ── Disk-based file cache (survives new browser sessions) ───────────────────
_CACHE_DIR = pathlib.Path(".file_cache")
_CACHE_DIR.mkdir(exist_ok=True)


def _save_to_disk(key: str, file_bytes: bytes, file_name: str) -> None:
    (_CACHE_DIR / f"{key}.bin").write_bytes(file_bytes)
    (_CACHE_DIR / f"{key}.name").write_text(file_name, encoding="utf-8")


def _load_from_disk(key: str):
    b_path = _CACHE_DIR / f"{key}.bin"
    n_path = _CACHE_DIR / f"{key}.name"
    if b_path.exists() and n_path.exists():
        return b_path.read_bytes(), n_path.read_text(encoding="utf-8")
    return None, None


def _clear_from_disk(key: str) -> None:
    for suffix in (".bin", ".name"):
        p = _CACHE_DIR / f"{key}{suffix}"
        if p.exists():
            p.unlink()


# ── Excel formatting helpers ──────────────────────────────────────────────────
def _style_ws(ws, pct_col="Compliance %", has_total_row=False):
    HDR_FILL   = PatternFill("solid", fgColor="2E75B6")
    ALT_FILL   = PatternFill("solid", fgColor="DCE6F1")
    HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
    GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
    AMBER_FILL = PatternFill("solid", fgColor="FFEB9C")
    RED_FILL   = PatternFill("solid", fgColor="FFC7CE")
    GREEN_FONT = Font(color="276221", bold=True)
    AMBER_FONT = Font(color="9C5700", bold=True)
    RED_FONT   = Font(color="9C0006", bold=True)
    max_row  = ws.max_row
    data_end = max_row - 1 if has_total_row else max_row
    large    = max_row > 400          # skip heavy per-cell loops for large sheets

    # Header — always applied
    for cell in ws[1]:
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    # Alternating row fill — skip for large sheets (too slow in openpyxl)
    if not large:
        for r in range(2, data_end + 1):
            for cell in ws[r]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if r % 2 == 0:
                    cell.fill = ALT_FILL

    # Column widths — sample first 200 rows only for large sheets
    scan_end = min(max_row, 200) if large else max_row
    for col in ws.iter_cols(min_row=1, max_row=scan_end):
        cl = get_column_letter(col[0].column)
        w  = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[cl].width = min(max(w + 3, 10), 42)

    ws.freeze_panes = "A2"

    # Compliance % conditional formatting
    pct_cl = None
    if pct_col:
        for cell in ws[1]:
            if cell.value == pct_col:
                pct_cl = get_column_letter(cell.column)
                break
        if pct_cl:
            rng = f"{pct_cl}2:{pct_cl}{data_end}"
            ws.conditional_formatting.add(
                rng, CellIsRule(operator="greaterThanOrEqual", formula=["0.7"],
                                fill=GREEN_FILL, font=GREEN_FONT))
            ws.conditional_formatting.add(
                rng, CellIsRule(operator="between", formula=["0.5", "0.6999"],
                                fill=AMBER_FILL, font=AMBER_FONT))
            ws.conditional_formatting.add(
                rng, CellIsRule(operator="lessThan", formula=["0.5"],
                                fill=RED_FILL, font=RED_FONT))
            for r in range(2, max_row + 1):
                ws[f"{pct_cl}{r}"].number_format = "0.0%"

    # Total row — always applied
    if has_total_row and max_row >= 2:
        for cell in ws[max_row]:
            cell.font      = HDR_FONT
            cell.fill      = HDR_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[max_row].height = 22


def _add_bar_chart(ws, name_col, title, exclude_last=False):
    hdr = [cell.value for cell in ws[1]]
    if "Compliance %" not in hdr or name_col not in hdr:
        return
    pct_idx  = hdr.index("Compliance %") + 1
    name_idx = hdr.index(name_col) + 1
    max_row  = ws.max_row - 1 if exclude_last else ws.max_row
    chart = BarChart()
    chart.type      = "bar"
    chart.grouping  = "clustered"
    chart.title     = title
    chart.style     = 2
    chart.height    = max(8, min(20, (max_row - 1) * 0.6 + 5))
    chart.width     = 18
    chart.y_axis.title = name_col
    chart.x_axis.title = "Compliance %"
    data_ref = Reference(ws, min_col=pct_idx, min_row=1, max_row=max_row)
    cats_ref = Reference(ws, min_col=name_idx, min_row=2, max_row=max_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    anchor = get_column_letter(ws.max_column + 2) + "2"
    ws.add_chart(chart, anchor)


WORKING_AUX: set[str] = {
    "outboundpending",
    "callbackpending",
    "consultpending",
    "inboundpending",
    "outbond",
    "outbound",
    "acw",
    "in call - working",
    "in call- working",
    "e-mails",
    "emails",
    "follow-up case work",
    "followup case work",
    "followup",
    "follow-up",
}


def is_working_aux(label: str) -> bool:
    """Return True if the AUX label represents active/working activity."""
    return label.lower().strip() in WORKING_AUX


@st.cache_data(show_spinner="Loading file…")
def load_file(file_bytes: bytes, file_name: str) -> pd.DataFrame:
    if file_name.lower().endswith(".csv"):
        return pd.read_csv(io.BytesIO(file_bytes))
    return pd.read_excel(io.BytesIO(file_bytes), engine="openpyxl")


@st.cache_data(show_spinner=False)
def get_excel_sheets(file_bytes: bytes) -> list[str]:
    """Return sheet names for an Excel file, or [] for CSV."""
    try:
        xl = pd.ExcelFile(io.BytesIO(file_bytes), engine="openpyxl")
        return xl.sheet_names
    except Exception:
        return []


@st.cache_data(show_spinner="Loading sheet…")
def load_file_sheet(file_bytes: bytes, file_name: str, sheet_name: str) -> pd.DataFrame:
    """Load a specific sheet from an Excel file (or the whole CSV)."""
    if file_name.lower().endswith(".csv"):
        return pd.read_csv(io.BytesIO(file_bytes))
    return pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, engine="openpyxl")


OT_SHEET_NAME = "Enrollment Register"
OT_SKIP_ROWS  = 3   # first 3 rows are metadata; row 4 is the real header


@st.cache_data(show_spinner="Loading OT schedule…")
def load_ot_schedule(file_bytes: bytes, file_name: str) -> pd.DataFrame:
    """Load the OT schedule from the 'Enrollment Register' sheet, skipping the first 3 rows."""
    if file_name.lower().endswith(".csv"):
        return pd.read_csv(io.BytesIO(file_bytes), skiprows=OT_SKIP_ROWS)
    sheets = get_excel_sheets(file_bytes)
    matched = next((s for s in sheets if s.strip().lower() == OT_SHEET_NAME.lower()), None)
    if matched is None:
        raise ValueError(
            f"Sheet '{OT_SHEET_NAME}' not found in the uploaded file. "
            f"Available sheets: {', '.join(sheets) if sheets else '(none)'}"
        )
    return pd.read_excel(
        io.BytesIO(file_bytes),
        sheet_name=matched,
        skiprows=OT_SKIP_ROWS,
        engine="openpyxl",
    )


def parse_datetimes(df: pd.DataFrame, login_col: str, logout_col: str) -> pd.DataFrame:
    df = df.copy()
    df[login_col]  = pd.to_datetime(df[login_col],  errors="coerce", dayfirst=False)
    df[logout_col] = pd.to_datetime(df[logout_col], errors="coerce", dayfirst=False)
    bad = df[login_col].isna() | df[logout_col].isna() | (df[login_col] > df[logout_col])
    if bad.any():
        st.sidebar.warning(f"⚠️ Skipping {bad.sum()} rows with invalid/missing datetimes.")
    df = df[~bad].reset_index(drop=True)
    return df


# ──────────────────────────────────────────────────────────────────────────────
# Sidebar – File upload & column mapping
# ──────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.header("📁 Data Source")
    uploaded = st.file_uploader("Upload Excel or CSV", type=["xlsx", "xls", "csv"])

    if uploaded is not None:
        _b = uploaded.read()
        _save_to_disk("login", _b, uploaded.name)
        st.session_state["login_bytes"] = _b
        st.session_state["login_name"]  = uploaded.name

    if "login_bytes" not in st.session_state:
        _b, _n = _load_from_disk("login")
        if _b is not None:
            st.session_state["login_bytes"] = _b
            st.session_state["login_name"]  = _n

    if "login_bytes" not in st.session_state:
        st.info("Upload a file to begin.")
        st.stop()

    st.caption(f"📄 **{st.session_state['login_name']}**")
    if st.button("🗑️ Clear login file", key="clear_login"):
        _clear_from_disk("login")
        st.session_state.pop("login_bytes", None)
        st.session_state.pop("login_name",  None)
        st.rerun()

    raw_df = load_file(st.session_state["login_bytes"], st.session_state["login_name"])
    st.success(f"✅ {len(raw_df):,} rows loaded")
    st.info("🕐 UTC → IST (+5:30) conversion applied")

    st.divider()
    st.subheader("🗂️ Column Mapping")
    st.caption("Auto-detected from common names – adjust if needed.")

    cols = list(raw_df.columns)

    def _idx(col): return cols.index(col) if col else 0

    agent_default  = find_col(raw_df, ["agent name", "agentname", "agent"])
    login_default  = find_col(raw_df, ["login date", "logindate", "start time", "starttime", "start date", "startdate"])
    logout_default = find_col(raw_df, ["logout date", "logoutdate", "end time", "endtime", "end date", "enddate"])
    aux_default    = find_col(raw_df, ["unavailable code", "unavailablecode", "aux code", "auxcode", "status code"])
    dur_default    = find_col(raw_df, ["duration"])

    agent_col  = st.selectbox("Agent Name",          cols, index=_idx(agent_default))
    login_col  = st.selectbox("Login / Start Time",  cols, index=_idx(login_default))
    logout_col = st.selectbox("Logout / End Time",   cols, index=_idx(logout_default))
    aux_col    = st.selectbox("AUX / Unavailable Code (optional)", ["— skip —"] + cols,
                               index=(_idx(aux_default) + 1) if aux_default else 0)
    if aux_col == "— skip —":
        aux_col = None

    skill_default = find_col(raw_df, ["skill name", "skillname", "skill", "queue name", "queue"])
    skill_col  = st.selectbox("Skill Name Column (optional)", ["— skip —"] + cols,
                               index=(_idx(skill_default) + 1) if skill_default else 0)
    if skill_col == "— skip —":
        skill_col = None

    avail_default = find_col(raw_df, ["available time", "availabletime", "available", "avail time", "avail"])
    avail_col = st.selectbox("Available Time Column (optional)", ["— skip —"] + cols,
                              index=(_idx(avail_default) + 1) if avail_default else 0)
    if avail_col == "— skip —":
        avail_col = None

# ──────────────────────────────────────────────────────────────────────────────
# Parse datetimes
# ──────────────────────────────────────────────────────────────────────────────
try:
    df = parse_datetimes(raw_df, login_col, logout_col)
except Exception as exc:
    st.error(f"Could not parse datetime columns: {exc}")
    st.stop()

if df.empty:
    st.error("No valid rows remain after parsing datetimes. Check your column mapping.")
    st.stop()

# ── UTC → IST conversion (UTC+5:30) ──────────────────────────────────────────
IST_OFFSET = timedelta(hours=5, minutes=30)
df[login_col]  = df[login_col]  + IST_OFFSET
df[logout_col] = df[logout_col] + IST_OFFSET

# ── Derive AUX label ─────────────────────────────────────────────────────────
def _is_nonzero(val):
    if pd.isna(val):
        return False
    s = str(val).strip()
    if not s or s.lower() in ("0", "nan", "-", "00:00:00", "00:00", "0:00:00"):
        return False
    try:
        return float(s) > 0
    except ValueError:
        pass
    try:
        return any(int(p) > 0 for p in s.split(":"))
    except Exception:
        return bool(s)


def derive_aux_label(row):
    aux_val   = str(row[aux_col]).strip()   if aux_col   and pd.notna(row[aux_col])   and str(row[aux_col]).strip()   != "" else ""
    skill_val = str(row[skill_col]).strip() if skill_col and pd.notna(row[skill_col]) and str(row[skill_col]).strip() != "" else ""
    if aux_val:
        return aux_val
    elif skill_val:
        return "In Call - Working"
    elif avail_col and _is_nonzero(row.get(avail_col)):
        return "Available"
    else:
        return "Unavailable - Not Working"

df["_aux_label"] = df.apply(derive_aux_label, axis=1)

# ──────────────────────────────────────────────────────────────────────────────
# Query controls
# ──────────────────────────────────────────────────────────────────────────────
st.markdown('<div class="section-header">🔍 Query Parameters</div>', unsafe_allow_html=True)
st.write("")

qc1, qc2, qc3, qc4 = st.columns([3, 2, 1.5, 1.5])

with qc1:
    agents = sorted(df[agent_col].dropna().unique().tolist())
    sel_agent = st.selectbox("👤 Agent", agents)

agent_df = df[df[agent_col] == sel_agent].copy()

with qc2:
    avail_dates = sorted(agent_df[login_col].dt.date.unique())
    sel_date = st.selectbox(
        "📅 Date (IST)",
        avail_dates,
        format_func=lambda d: d.strftime("%a, %b %d %Y"),
    )

with qc3:
    slot_start = st.time_input("⏰ From (IST)", value=time(13, 0))

with qc4:
    slot_end = st.time_input("⏰ To (IST)", value=time(14, 0))

if slot_start == slot_end:
    st.error("'From' and 'To' times cannot be the same.")
    st.stop()

# ──────────────────────────────────────────────────────────────────────────────
# Overlap filtering logic
# ──────────────────────────────────────────────────────────────────────────────
q_start = datetime.combine(sel_date, slot_start)
q_end   = datetime.combine(sel_date, slot_end)
if q_end <= q_start:            # midnight-crossing slot (e.g. 11 PM → 12 AM)
    q_end += timedelta(days=1)

# All rows for that agent on that date
day_df = agent_df[agent_df[login_col].dt.date == sel_date].copy()

# Overlap condition: row interval intersects [q_start, q_end]
overlap_mask = (day_df[login_col] < q_end) & (day_df[logout_col] > q_start)
result_df = day_df[overlap_mask].copy()

if not result_df.empty:
    # Precise logout = login + Duration when Duration column has sub-minute accuracy
    if dur_default and dur_default in result_df.columns:
        _dur_s = result_df[dur_default].apply(_parse_dur_secs)
        _has_dur = _dur_s.notna() & (_dur_s > 0)
        _precise_end = result_df[logout_col].copy()
        _precise_end[_has_dur] = (
            result_df.loc[_has_dur, login_col]
            + _dur_s[_has_dur].apply(lambda s: timedelta(seconds=s))
        )
    else:
        _precise_end = result_df[logout_col]

    # Effective (clipped) times within the queried slot
    result_df["_eff_start"] = result_df[login_col].clip(lower=q_start, upper=q_end)
    result_df["_eff_end"]   = _precise_end.clip(lower=q_start, upper=q_end)
    result_df["_overlap_sec"] = (
        result_df["_eff_end"] - result_df["_eff_start"]
    ).dt.total_seconds()
    result_df["_overlap_min"] = result_df["_overlap_sec"] / 60

# ──────────────────────────────────────────────────────────────────────────────
# Display results
# ──────────────────────────────────────────────────────────────────────────────
st.divider()
st.markdown(
    f'<div class="section-header">📋 Results — {sel_agent}</div>',
    unsafe_allow_html=True,
)
st.caption(
    f"Time slot (IST): **{slot_start.strftime('%H:%M')}** → **{slot_end.strftime('%H:%M')}** "
    f"on **{sel_date.strftime('%A, %B %d, %Y')}**"
)

if result_df.empty:
    st.warning("No AUX records overlap the selected time slot for this agent on this date.")
else:
    slot_total_sec = (q_end - q_start).total_seconds()
    # Union of all record intervals (not simple sum) to avoid double-counting overlaps
    total_overlap_sec = _union_secs(
        list(zip(result_df["_eff_start"], result_df["_eff_end"]))
    )

    # ── Metrics ──────────────────────────────────────────────────────────────
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Matching Records", len(result_df))
    m2.metric("Total AUX Time in Slot", fmt_duration(total_overlap_sec))
    m3.metric("Slot Duration", fmt_duration(slot_total_sec))
    m4.metric("Slot Coverage", f"{total_overlap_sec / slot_total_sec * 100:.1f}%")

    st.write("")

    # ── Detailed table ────────────────────────────────────────────────────────
    st.markdown("#### 📄 Detailed Records")

    show_df = result_df.drop(columns=["_eff_start", "_eff_end", "_overlap_sec", "_overlap_min"]).copy()
    show_df = show_df.rename(columns={"_aux_label": "AUX Label"})

    # Add human-readable overlap columns
    show_df.insert(
        show_df.columns.get_loc(login_col) + 1,
        "Slot Effective Start (IST)",
        result_df["_eff_start"].dt.strftime("%H:%M:%S"),
    )
    show_df.insert(
        show_df.columns.get_loc(logout_col) + 1,
        "Slot Effective End (IST)",
        result_df["_eff_end"].dt.strftime("%H:%M:%S"),
    )
    show_df["Time in Slot (HH:MM:SS)"] = result_df["_overlap_sec"].apply(fmt_duration)

    # Format datetime cols for readability and rename to show IST
    for dc in [login_col, logout_col]:
        show_df[dc] = show_df[dc].dt.strftime("%m/%d/%Y %H:%M:%S")
    show_df = show_df.rename(columns={
        login_col:  f"{login_col} (IST)",
        logout_col: f"{logout_col} (IST)",
    })

    st.dataframe(show_df, use_container_width=True, hide_index=True)

    # ── AUX Breakdown ─────────────────────────────────────────────────────────
    if aux_col or skill_col:
        st.write("")
        st.markdown("#### 📊 AUX Breakdown in Slot")

        _triples = list(zip(
            result_df["_aux_label"],
            result_df["_eff_start"],
            result_df["_eff_end"],
        ))
        _alloc = _allocate_aux_secs(_triples)
        summary = (
            pd.DataFrame([{"AUX Code": k, "Seconds": v} for k, v in _alloc.items()])
            if _alloc else pd.DataFrame(columns=["AUX Code", "Seconds"])
        )
        summary = summary.sort_values("Seconds", ascending=False).reset_index(drop=True)
        summary["Duration (HH:MM:SS)"] = summary["Seconds"].apply(fmt_duration)
        summary["Minutes"] = (summary["Seconds"] / 60).round(2)
        summary["% of Slot"] = (summary["Seconds"] / slot_total_sec * 100).round(1)

        sc1, sc2 = st.columns([1, 1.5])

        with sc1:
            st.dataframe(
                summary[["AUX Code", "Duration (HH:MM:SS)", "% of Slot"]],
                use_container_width=True,
                hide_index=True,
            )

        with sc2:
            pie_data = summary[summary["Seconds"] > 0]
            if not pie_data.empty:
                fig_pie = px.pie(
                    pie_data,
                    values="Seconds",
                    names="AUX Code",
                    title=f"AUX Distribution  {slot_start.strftime('%H:%M')}–{slot_end.strftime('%H:%M')} IST",
                    hole=0.42,
                    color_discrete_sequence=px.colors.qualitative.Set2,
                )
                fig_pie.update_traces(textinfo="label+percent", pull=[0.03] * len(pie_data))
                fig_pie.update_layout(
                    height=320,
                    margin=dict(t=40, b=10, l=10, r=10),
                    showlegend=False,
                )
                st.plotly_chart(fig_pie, use_container_width=True)

    # ── Timeline / Gantt ──────────────────────────────────────────────────────
    st.write("")
    st.markdown("#### ⏱️ Activity Timeline (within slot)")

    timeline_rows = []
    for _, row in result_df.iterrows():
        label = row["_aux_label"]
        timeline_rows.append(
            {
                "AUX": label,
                "Start": row["_eff_start"],
                "Finish": row["_eff_end"],
            }
        )

    tl_df = pd.DataFrame(timeline_rows)

    if not tl_df.empty:
        # Sort so stacking looks sensible
        tl_df = tl_df.sort_values("Start")
        fig_gantt = px.timeline(
            tl_df,
            x_start="Start",
            x_end="Finish",
            y="AUX",
            color="AUX",
            color_discrete_sequence=px.colors.qualitative.Set2,
            title=f"Agent Timeline: {sel_agent}  (IST)",
        )
        fig_gantt.update_xaxes(
            range=[q_start, q_end],
            tickformat="%H:%M",
            title="Time (IST)",
        )
        fig_gantt.update_yaxes(title="")
        fig_gantt.update_layout(
            height=max(280, 60 + 45 * tl_df["AUX"].nunique()),
            showlegend=False,
            margin=dict(t=40, b=30, l=10, r=10),
        )
        st.plotly_chart(fig_gantt, use_container_width=True)

    # ── Export ────────────────────────────────────────────────────────────────
    st.write("")
    st.markdown("#### 💾 Export")

    out_buf = io.BytesIO()
    _show_clean = show_df.loc[:, ~show_df.columns.str.startswith("Unnamed:")].copy()
    with pd.ExcelWriter(out_buf, engine="openpyxl") as writer:
        _show_clean.to_excel(writer, sheet_name="Filtered Records", index=False)
        _style_ws(writer.sheets["Filtered Records"], pct_col=None)
        if aux_col or skill_col:
            summary[["AUX Code", "Duration (HH:MM:SS)", "Minutes", "% of Slot"]].to_excel(
                writer, sheet_name="AUX Summary", index=False
            )
            _style_ws(writer.sheets["AUX Summary"], pct_col=None)

    file_label = (
        f"AUX_{sel_agent.replace(' ', '_').replace(',', '')}_{sel_date}"
        f"_{slot_start.strftime('%H%M')}-{slot_end.strftime('%H%M')}.xlsx"
    )
    st.download_button(
        "📥 Download Results (Excel)",
        data=out_buf.getvalue(),
        file_name=file_label,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# ──────────────────────────────────────────────────────────────────────────────
# Full-day context (collapsible)
# ──────────────────────────────────────────────────────────────────────────────
with st.expander(f"🗓️ Full Day View — {sel_agent}  ({sel_date.strftime('%b %d, %Y')})", expanded=False):
    if day_df.empty:
        st.write("No records for this agent on the selected date.")
    else:
        day_show = day_df.copy()
        for dc in [login_col, logout_col]:
            day_show[dc] = day_show[dc].dt.strftime("%m/%d/%Y %H:%M:%S")
        day_show = day_show.rename(columns={
            login_col:    f"{login_col} (IST)",
            logout_col:   f"{logout_col} (IST)",
            "_aux_label": "AUX Label",
        })
        st.dataframe(day_show, use_container_width=True, hide_index=True)

        # Full-day timeline
        if aux_col or skill_col:
            fd_rows = []
            for _, row in day_df.iterrows():
                label = row["_aux_label"]
                fd_rows.append({"AUX": label, "Start": row[login_col], "Finish": row[logout_col]})

            fd_tl = pd.DataFrame(fd_rows).sort_values("Start")
            if not fd_tl.empty:
                fig_fd = px.timeline(
                    fd_tl,
                    x_start="Start",
                    x_end="Finish",
                    y="AUX",
                    color="AUX",
                    color_discrete_sequence=px.colors.qualitative.Set2,
                    title="Full Day Activity (IST)",
                )
                fig_fd.update_xaxes(tickformat="%H:%M", title="Time (IST)")
                fig_fd.update_yaxes(title="")
                fig_fd.update_layout(
                    height=max(300, 60 + 45 * fd_tl["AUX"].nunique()),
                    showlegend=False,
                    margin=dict(t=40, b=30, l=10, r=10),
                )
                # Highlight the queried slot with a shaded rectangle
                fig_fd.add_vrect(
                    x0=q_start, x1=q_end,
                    fillcolor="rgba(21, 101, 192, 0.12)",
                    line_width=1,
                    line_color="rgba(21, 101, 192, 0.6)",
                    annotation_text=f"Queried slot",
                    annotation_position="top left",
                )
                st.plotly_chart(fig_fd, use_container_width=True)

# ──────────────────────────────────────────────────────────────────────────────
# Schedule Compliance Report
# ──────────────────────────────────────────────────────────────────────────────
st.divider()
st.markdown('<div class="section-header">📋 Schedule Compliance Report</div>', unsafe_allow_html=True)
st.write("")
st.caption(
    "Upload the agent schedule file (Excel/CSV). "
    "Agent names in the schedule must **exactly match** the Agent Name values in the login data. "
    "Expected day columns: Sun, Mon, Tue, Wed, Thu, Fri, Sat — values like '6 PM-7PM', "
    "'10 PM to 11 PM', or 'WO' (week off)."
)

sched_file = st.file_uploader(
    "Upload Schedule File (Excel / CSV)",
    type=["xlsx", "xls", "csv"],
    key="sched_upload",
)

if sched_file is not None:
    _sb = sched_file.read()
    _save_to_disk("sched", _sb, sched_file.name)
    st.session_state["sched_bytes"] = _sb
    st.session_state["sched_name"]  = sched_file.name

if "sched_bytes" not in st.session_state:
    _sb, _sn = _load_from_disk("sched")
    if _sb is not None:
        st.session_state["sched_bytes"] = _sb
        st.session_state["sched_name"]  = _sn

if "sched_bytes" in st.session_state:
    sc1, sc2 = st.columns([1, 3])
    if sc1.button("🗑️ Clear schedule file", key="clear_sched"):
        _clear_from_disk("sched")
        st.session_state.pop("sched_bytes", None)
        st.session_state.pop("sched_name",  None)
        st.rerun()
    sc2.caption(f"📄 Using: **{st.session_state['sched_name']}**")

    sched_raw = load_file(st.session_state["sched_bytes"], st.session_state["sched_name"])
    st.success(f"Schedule loaded — {len(sched_raw)} agent rows")

    with st.expander("Preview Schedule Data", expanded=False):
        st.dataframe(sched_raw, use_container_width=True)

    st.markdown("##### Configure Schedule Columns")
    sched_cols = list(sched_raw.columns)
    cfg1, cfg2, cfg3 = st.columns([2, 2, 3])

    with cfg1:
        sched_agent_default = find_col(sched_raw, ["agent name", "agentname", "agent"])
        sched_agent_col = st.selectbox(
            "Agent Name column (in schedule)",
            sched_cols,
            index=sched_cols.index(sched_agent_default) if sched_agent_default else 0,
            key="sched_agent_col",
        )

    with cfg2:
        sched_sup_default = find_col(sched_raw, [
            "supervisor name", "supervisorname", "supervisor",
            "team lead", "teamlead", "tl", "manager", "reporting to", "reports to",
        ])
        _sup_opts = ["— skip —"] + sched_cols
        sched_sup_col = st.selectbox(
            "Supervisor column (optional)",
            _sup_opts,
            index=(_sup_opts.index(sched_sup_default)
                   if sched_sup_default and sched_sup_default in _sup_opts else 0),
            key="sched_sup_col",
        )
        if sched_sup_col == "— skip —":
            sched_sup_col = None

    with cfg3:
        auto_days = [
            c for c in sched_cols
            if c.strip()[:3].lower() in ("sun", "mon", "tue", "wed", "thu", "fri", "sat")
        ]
        day_cols_sel = st.multiselect(
            "Day-of-week columns (select all that apply)",
            sched_cols,
            default=auto_days,
            key="day_cols_sel",
        )

    # Date range to analyse
    all_dates = sorted(df[login_col].dt.date.unique())
    dr1, dr2 = st.columns(2)
    with dr1:
        date_from = st.date_input("From Date (IST)", value=min(all_dates), key="sched_from")
    with dr2:
        date_to = st.date_input("To Date (IST)",   value=max(all_dates), key="sched_to")

    check_dates = [d for d in all_dates if date_from <= d <= date_to]

    if not check_dates:
        st.warning("No dates in the login data fall within the selected range.")
    else:
        st.info(
            f"Checking **{len(check_dates)} date(s)** × "
            f"**{len(sched_raw)} agent row(s)** = "
            f"**{len(check_dates) * len(sched_raw)} slot checks**."
        )

        if st.button("📊 Generate Compliance Report", type="primary"):

            # Build day-of-week abbreviation → schedule column name
            _dow_norm = {
                "sun": "Sun", "mon": "Mon", "tue": "Tue",
                "wed": "Wed", "thu": "Thu", "fri": "Fri", "sat": "Sat",
            }
            dow_col_map = {}
            for col in day_cols_sel:
                prefix = col.strip()[:3].lower()
                if prefix in _dow_norm:
                    dow_col_map[_dow_norm[prefix]] = col

            # weekday() → English day abbreviation (locale-independent)
            _wd_to_dow = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

            records = []
            out_records = []
            aux_detail_records = []
            total_ops = max(len(sched_raw), 1)
            prog = st.progress(0.0, text="Generating report…")
            done = 0

            # Pre-compute login date and group by (agent, date) for O(1) lookup
            df["_login_date"] = df[login_col].dt.date
            _empty_login_df  = df.iloc[0:0]
            _login_lookup    = {
                key: grp
                for key, grp in df.groupby([agent_col, "_login_date"], sort=False)
            }

            for _, sched_row in sched_raw.iterrows():
                sched_agent = str(sched_row[sched_agent_col]).strip()
                if not sched_agent or sched_agent.lower() == "nan":
                    done += 1
                    prog.progress(min(done / total_ops, 1.0))
                    continue

                sched_supervisor = (
                    str(sched_row[sched_sup_col]).strip()
                    if sched_sup_col and pd.notna(sched_row.get(sched_sup_col))
                    else ""
                )
                for check_date in check_dates:
                    dow_short = _wd_to_dow[check_date.weekday()]

                    slot_str = "WO"
                    if dow_short in dow_col_map:
                        raw = sched_row.get(dow_col_map[dow_short], "WO")
                        slot_str = str(raw).strip() if pd.notna(raw) else "WO"

                    base = {
                        "Supervisor":     sched_supervisor,
                        "Agent Name":     sched_agent,
                        "Date (IST)":     check_date.strftime("%d-%b-%Y"),
                        "Day":            dow_short,
                        "Scheduled Slot": slot_str,
                    }

                    if slot_str.upper() in ("WO", "", "NAN", "NONE", "-"):
                        records.append({
                            **base,
                            "Scheduled Slot":       "Week Off",
                            "Status":               "📅 Week Off",
                            "In Call Duration":     "-",
                            "Total Active in Slot": "-",
                            "AUX Breakdown":        "-",
                            "Inbound Handled":      0,
                            "Outbound Handled":     0,
                        })
                        wo_rows = _login_lookup.get((sched_agent, check_date), _empty_login_df)
                        for _, lrow in wo_rows.iterrows():
                            if not is_working_aux(lrow["_aux_label"]):
                                continue
                            _ds = max(0, (lrow[logout_col] - lrow[login_col]).total_seconds())
                            if _ds > 0:
                                out_records.append({
                                    "Supervisor":          sched_supervisor,
                                    "Agent Name":          sched_agent,
                                    "Date (IST)":          check_date.strftime("%d-%b-%Y"),
                                    "Day":                 dow_short,
                                    "Scheduled Slot":      "Week Off",
                                    "Login Time (IST)":    lrow[login_col].strftime("%H:%M"),
                                    "Logout Time (IST)":   lrow[logout_col].strftime("%H:%M"),
                                    "AUX Code":            lrow["_aux_label"],
                                    "Duration":            fmt_duration(_ds),
                                    "Type":                "Week Off - Logged In",
                                })

                    else:
                        parsed = parse_time_slot(slot_str)

                        if parsed is None:
                            records.append({
                                **base,
                                "Status":               "⚠️ Slot Parse Error",
                                "In Call Duration":     "-",
                                "Total Active in Slot": "-",
                                "AUX Breakdown":        f"Cannot parse: '{slot_str}'",
                                "Inbound Handled":      0,
                                "Outbound Handled":     0,
                            })

                        else:
                            s_time, e_time = parsed
                            q_s = datetime.combine(check_date, s_time)
                            q_e = datetime.combine(check_date, e_time)
                            if q_e <= q_s:    # midnight-crossing slot
                                q_e += timedelta(days=1)

                            day_rows = _login_lookup.get((sched_agent, check_date), _empty_login_df)
                            overlap  = day_rows[
                                (day_rows[login_col] < q_e) & (day_rows[logout_col] > q_s)
                            ].copy()

                            for _, lrow in day_rows.iterrows():
                                if not is_working_aux(lrow["_aux_label"]):
                                    continue
                                _lt  = lrow[login_col]
                                _lo  = lrow[logout_col]
                                _tot = max(0, (_lo - _lt).total_seconds())
                                _es  = max(_lt, q_s)
                                _ee  = min(_lo, q_e)
                                _in  = max(0, (_ee - _es).total_seconds()) if _ee > _es else 0
                                _out = _tot - _in
                                if _out > 0:
                                    if _lt < q_s and _lo <= q_s:
                                        _otype = "Before Slot"
                                    elif _lt >= q_e:
                                        _otype = "After Slot"
                                    elif _lt < q_s and _lo > q_e:
                                        _otype = "Before & After Slot"
                                    elif _lt < q_s:
                                        _otype = "Partially Before Slot"
                                    else:
                                        _otype = "Partially After Slot"
                                    out_records.append({
                                        "Supervisor":           sched_supervisor,
                                        "Agent Name":           sched_agent,
                                        "Date (IST)":           check_date.strftime("%d-%b-%Y"),
                                        "Day":                  dow_short,
                                        "Scheduled Slot":       slot_str,
                                        "Login Time (IST)":     _lt.strftime("%H:%M:%S"),
                                        "Logout Time (IST)":    _lo.strftime("%H:%M:%S"),
                                        "AUX Code":             lrow["_aux_label"],
                                        "Out-of-Slot Duration": fmt_duration(_out),
                                        "Type":                 _otype,
                                    })

                            if overlap.empty:
                                records.append({
                                    **base,
                                    "Status":               "❌ Not Logged In",
                                    "In Call Duration":     "00:00:00",
                                    "Total Active in Slot": "00:00:00",
                                    "AUX Breakdown":        "No activity",
                                    "Inbound Handled":      0,
                                    "Outbound Handled":     0,
                                })
                            else:
                                # Precise logout = login + Duration for sub-minute accuracy
                                if dur_default and dur_default in overlap.columns:
                                    _d_s = overlap[dur_default].apply(_parse_dur_secs)
                                    _hd  = _d_s.notna() & (_d_s > 0)
                                    _prec_end = overlap[logout_col].copy()
                                    _prec_end[_hd] = (
                                        overlap.loc[_hd, login_col]
                                        + _d_s[_hd].apply(lambda s: timedelta(seconds=s))
                                    )
                                else:
                                    _prec_end = overlap[logout_col]

                                overlap["_es"] = overlap[login_col].clip(lower=q_s, upper=q_e)
                                overlap["_ee"] = _prec_end.clip(lower=q_s, upper=q_e)
                                overlap["_sc"] = (overlap["_ee"] - overlap["_es"]).dt.total_seconds()

                                # ── Interval-union totals (no double-counting) ──────────
                                _all_ivs = list(zip(overlap["_es"], overlap["_ee"]))
                                total_secs = _union_secs(_all_ivs)

                                # Working AUX intervals
                                _wk_mask = overlap["_aux_label"].apply(is_working_aux)
                                _wk_rows = overlap[_wk_mask]
                                _wk_ivs  = list(zip(_wk_rows["_es"], _wk_rows["_ee"]))

                                in_call_secs = _union_secs(_wk_ivs)

                                _cpl_triples = list(zip(
                                    overlap["_aux_label"],
                                    overlap["_es"],
                                    overlap["_ee"],
                                ))
                                aux_grp = pd.Series(
                                    _allocate_aux_secs(_cpl_triples)
                                ).sort_values(ascending=False)
                                aux_str = "; ".join(
                                    f"{k}: {fmt_duration(v)}"
                                    for k, v in aux_grp.items() if v > 0
                                )
                                for _albl, _asecs in aux_grp.items():
                                    if is_working_aux(_albl) and _asecs > 0:
                                        aux_detail_records.append({
                                            "Supervisor": sched_supervisor,
                                            "AUX Code":   _albl,
                                            "_secs":      _asecs,
                                        })

                                if aux_col:
                                    _aux_blank = (
                                        overlap[aux_col].isna() |
                                        (overlap[aux_col].astype(str).str.strip() == "")
                                    )
                                else:
                                    _aux_blank = pd.Series(True, index=overlap.index)
                                if skill_col:
                                    _skill_pop = (
                                        overlap[skill_col].notna() &
                                        (overlap[skill_col].astype(str).str.strip() != "")
                                    )
                                    if dur_default and dur_default in overlap.columns:
                                        _dur_s  = overlap[dur_default].fillna("").astype(str).str.strip().str.lower()
                                        _dur_ok = ~_dur_s.isin(["", "0", "nan", "none", "-", "00:00:00", "00:00", "0:00:00"])
                                    else:
                                        _dur_ok = pd.Series(True, index=overlap.index)
                                    _call_mask = _aux_blank & _skill_pop & _dur_ok
                                    _outbound_cnt = int(
                                        overlap.loc[_call_mask, skill_col]
                                        .fillna("").str.lower()
                                        .str.contains("outbond|outbound", na=False).sum()
                                    )
                                    _inbound_cnt = int(_call_mask.sum()) - _outbound_cnt
                                else:
                                    _inbound_cnt = _outbound_cnt = 0

                                status = (
                                    "✅ Logged In - On Calls"
                                    if in_call_secs > 0
                                    else "⚠️ Logged In - No Calls"
                                )
                                records.append({
                                    **base,
                                    "Status":               status,
                                    "In Call Duration":     fmt_duration(in_call_secs),
                                    "Total Active in Slot": fmt_duration(total_secs),
                                    "AUX Breakdown":        aux_str,
                                    "Inbound Handled":      _inbound_cnt,
                                    "Outbound Handled":     _outbound_cnt,
                                })

                done += 1
                prog.progress(min(done / total_ops, 1.0))

            prog.empty()

            if not records:
                st.info(
                    "No records generated. "
                    "Verify that agent names in the schedule match the login data exactly."
                )
            else:
                report_df = pd.DataFrame(records)

                # ── Summary metrics ───────────────────────────────────────────
                non_wo    = report_df[~report_df["Status"].str.startswith("📅", na=False)]
                on_calls  = non_wo[non_wo["Status"].str.startswith("✅", na=False)]
                no_calls  = non_wo[non_wo["Status"].str.startswith("⚠️", na=False)]
                not_in    = non_wo[non_wo["Status"].str.startswith("❌", na=False)]
                week_off  = report_df[report_df["Status"].str.startswith("📅", na=False)]

                sm1, sm2, sm3, sm4, sm5 = st.columns(5)
                sm1.metric("Scheduled Slots",       len(non_wo))
                sm2.metric("✅ On Calls",             len(on_calls))
                sm3.metric("⚠️ Logged - No Calls",   len(no_calls))
                sm4.metric("❌ Not Logged In",        len(not_in))
                sm5.metric("📅 Week Off",             len(week_off))

                # ── Compliance % per agent ────────────────────────────────────
                if len(non_wo) > 0:
                    _stats = []
                    _ag_grp_cols = (
                        ["Supervisor", "Agent Name"]
                        if sched_sup_col and "Supervisor" in non_wo.columns
                        else ["Agent Name"]
                    )
                    for _keys_ag, _grp in non_wo.groupby(_ag_grp_cols):
                        if not isinstance(_keys_ag, tuple):
                            _keys_ag = (_keys_ag,)
                        _tot = len(_grp)
                        _on  = int(_grp["Status"].str.startswith("✅").sum())
                        _nc  = int(_grp["Status"].str.startswith("⚠️").sum())
                        _ni  = int(_grp["Status"].str.startswith("❌").sum())
                        _stats.append({
                            **dict(zip(_ag_grp_cols, _keys_ag)),
                            "Scheduled Days":  _tot,
                            "✅ On Calls":      _on,
                            "⚠️ No Calls":      _nc,
                            "❌ Not Logged In": _ni,
                            "Compliance %":    round(_on / _tot, 4) if _tot > 0 else 0.0,
                        })
                    agent_summary = pd.DataFrame(_stats)

                    st.markdown("##### Agent Summary")
                    _ag_disp = agent_summary.copy()
                    _ag_disp["Compliance %"] = _ag_disp["Compliance %"].apply(lambda x: f"{x:.1%}")
                    st.dataframe(_ag_disp, use_container_width=True, hide_index=True)

                # ── Supervisor summary ───────────────────────────────────────────
                sup_summary = pd.DataFrame()
                if sched_sup_col and len(non_wo) > 0 and "Supervisor" in non_wo.columns:
                    _ss = []
                    for _sup, _sg in non_wo.groupby("Supervisor"):
                        _stot = len(_sg)
                        _son  = int(_sg["Status"].str.startswith("✅").sum())
                        _snc  = int(_sg["Status"].str.startswith("⚠️").sum())
                        _sni  = int(_sg["Status"].str.startswith("❌").sum())
                        _ss.append({
                            "Supervisor":      _sup,
                            "Agents":          _sg["Agent Name"].nunique(),
                            "Scheduled Slots": _stot,
                            "✅ On Calls":      _son,
                            "⚠️ No Calls":      _snc,
                            "❌ Not Logged In": _sni,
                            "Compliance %":    round(_son / _stot, 4) if _stot > 0 else 0.0,
                        })
                    sup_summary = pd.DataFrame(_ss)
                    st.markdown("##### 👔 Supervisor Summary")
                    _sup_disp = sup_summary.copy()
                    _sup_disp["Compliance %"] = _sup_disp["Compliance %"].apply(lambda x: f"{x:.1%}")
                    st.dataframe(_sup_disp, use_container_width=True, hide_index=True)

                # ── Overall Summary Pivot ─────────────────────────────────────────
                st.markdown("##### 📊 Overall Summary Pivot")
                st.caption(
                    "Supervisor → Agent Name × Date — each date has two columns: "
                    "**In Call Duration** (within scheduled slot) and "
                    "**Out-of-Slot Duration** (outside scheduled slot / WO day)."
                )

                def _parse_hms(s):
                    try:
                        if str(s).strip() in ("-", "", "nan"):
                            return 0
                        _h, _m, _s = str(s).strip().split(":")
                        return int(_h) * 3600 + int(_m) * 60 + int(_s)
                    except Exception:
                        return 0

                _pdata2 = report_df[
                    ~report_df["Status"].str.startswith("📅", na=False)
                ].copy()
                _pdata2["_in_secs"] = _pdata2["In Call Duration"].apply(_parse_hms)

                _idx_cols_piv = (
                    ["Supervisor", "Agent Name"]
                    if sched_sup_col and "Supervisor" in _pdata2.columns
                    else ["Agent Name"]
                )

                _in_agg = (
                    _pdata2
                    .groupby(_idx_cols_piv + ["Date (IST)"])["_in_secs"]
                    .sum()
                    .reset_index()
                )

                if out_records:
                    _oos_df2 = pd.DataFrame(out_records).copy()
                    _oos_dur = _oos_df2.get(
                        "Out-of-Slot Duration",
                        pd.Series("00:00:00", index=_oos_df2.index),
                    ).fillna(
                        _oos_df2.get("Duration",
                                     pd.Series("00:00:00", index=_oos_df2.index))
                    ).fillna("00:00:00")
                    _oos_df2["_oos_secs"] = _oos_dur.apply(_parse_hms)
                    _oos_cols2 = [c for c in _idx_cols_piv if c in _oos_df2.columns]
                    _oos_agg = (
                        _oos_df2
                        .groupby(_oos_cols2 + ["Date (IST)"])["_oos_secs"]
                        .sum()
                        .reset_index()
                    )
                    for _c2 in _idx_cols_piv:
                        if _c2 not in _oos_agg.columns:
                            _oos_agg[_c2] = ""
                else:
                    _oos_agg = pd.DataFrame(
                        columns=_idx_cols_piv + ["Date (IST)", "_oos_secs"]
                    )

                _clong = (
                    _in_agg
                    .merge(_oos_agg, on=_idx_cols_piv + ["Date (IST)"], how="outer")
                    .fillna(0)
                )
                if "_oos_secs" not in _clong.columns:
                    _clong["_oos_secs"] = 0

                _piv_dates2 = sorted(
                    _clong["Date (IST)"].unique(),
                    key=lambda _d: datetime.strptime(_d, "%d-%b-%Y"),
                )

                _piv_rows = []
                for _key, _grp in _clong.groupby(_idx_cols_piv, sort=True):
                    if not isinstance(_key, tuple):
                        _key = (_key,)
                    _row = dict(zip(_idx_cols_piv, _key))
                    _tot_in = _tot_oos = 0
                    for _date in _piv_dates2:
                        _dg    = _grp[_grp["Date (IST)"] == _date]
                        _in_s  = int(_dg["_in_secs"].sum())
                        _oos_s = int(_dg["_oos_secs"].sum())
                        _row[f"{_date} | In Call"]     = fmt_duration(_in_s)
                        _row[f"{_date} | Out-of-Slot"] = fmt_duration(_oos_s)
                        _tot_in  += _in_s
                        _tot_oos += _oos_s
                    _row["TOTAL | In Call"]     = fmt_duration(_tot_in)
                    _row["TOTAL | Out-of-Slot"] = fmt_duration(_tot_oos)
                    _piv_rows.append(_row)

                _piv_combined = (
                    pd.DataFrame(_piv_rows) if _piv_rows else pd.DataFrame()
                )
                st.dataframe(_piv_combined, use_container_width=True, hide_index=True)

                # ── Hrs Logged vs Committed Hrs Pivot ─────────────────────────────
                st.markdown("##### ⏱️ Hrs Logged vs Committed Hrs Pivot")
                st.caption(
                    "For each date: **Count of Scheduled Slots** and **Sum of In Call Duration** (all working AUX). "
                    "**Compliance %** = Total In Call Duration ÷ Total Committed (scheduled slot) hours."
                )

                def _slot_secs(slot_str):
                    try:
                        _p = parse_time_slot(str(slot_str).strip())
                        if _p is None:
                            return 0
                        _dummy = date(2000, 1, 1)
                        _s_dt  = datetime.combine(_dummy, _p[0])
                        _e_dt  = datetime.combine(_dummy, _p[1])
                        if _e_dt <= _s_dt:      # midnight-crossing slot (e.g. 11 PM → 12 AM)
                            _e_dt += timedelta(days=1)
                        return (_e_dt - _s_dt).total_seconds()
                    except Exception:
                        return 0

                _pdata3 = report_df[
                    ~report_df["Status"].str.startswith("📅", na=False)
                ].copy()
                _pdata3["_in3_secs"]       = _pdata3["In Call Duration"].apply(_parse_hms)
                _pdata3["_committed_secs"] = _pdata3["Scheduled Slot"].apply(_slot_secs)

                _grp3 = (
                    _pdata3
                    .groupby(_idx_cols_piv + ["Date (IST)"], sort=True)
                    .agg(
                        _slot_cnt       = ("Scheduled Slot",  "count"),
                        _in3_secs       = ("_in3_secs",       "sum"),
                        _committed_secs = ("_committed_secs", "sum"),
                        _inbound        = ("Inbound Handled",  "sum"),
                        _outbound       = ("Outbound Handled", "sum"),
                    )
                    .reset_index()
                )

                _piv3_dates = sorted(
                    _grp3["Date (IST)"].unique(),
                    key=lambda _d: datetime.strptime(_d, "%d-%b-%Y"),
                )

                _piv3_rows = []
                for _key3, _g3 in _grp3.groupby(_idx_cols_piv, sort=True):
                    if not isinstance(_key3, tuple):
                        _key3 = (_key3,)
                    _r3 = dict(zip(_idx_cols_piv, _key3))
                    _ttl_cnt = _ttl_in = _ttl_comm = _ttl_inb = _ttl_outb = 0
                    for _d3 in _piv3_dates:
                        _dg3   = _g3[_g3["Date (IST)"] == _d3]
                        _cnt3  = int(_dg3["_slot_cnt"].sum())
                        _in3   = int(_dg3["_in3_secs"].sum())
                        _comm3 = int(_dg3["_committed_secs"].sum())
                        _r3[f"{_d3} | Count of Scheduled Slot"] = _cnt3
                        _r3[f"{_d3} | Scheduled Hrs"]           = fmt_duration(_comm3)
                        _r3[f"{_d3} | Sum of In Call Duration"]  = fmt_duration(_in3)
                        _ttl_cnt  += _cnt3
                        _ttl_in   += _in3
                        _ttl_comm += _comm3
                        _ttl_inb  += int(_dg3["_inbound"].sum())
                        _ttl_outb += int(_dg3["_outbound"].sum())
                    _r3["Total Count of Scheduled Slot"] = _ttl_cnt
                    _r3["Total Scheduled Hrs"]           = fmt_duration(_ttl_comm)
                    _r3["Total Sum of In Call Duration"]  = fmt_duration(_ttl_in)
                    _r3["Compliance %"] = (
                        _ttl_in / _ttl_comm
                        if _ttl_comm > 0 else 0.0
                    )
                    _r3["Inbound Handled"]  = _ttl_inb
                    _r3["Outbound Handled"] = _ttl_outb
                    _piv3_rows.append(_r3)

                _piv3 = pd.DataFrame(_piv3_rows) if _piv3_rows else pd.DataFrame()
                st.dataframe(_piv3, use_container_width=True, hide_index=True)

                # ── Supervisor-level Hrs vs Committed Pivot ────────────────────
                if sched_sup_col and "Supervisor" in _pdata3.columns:
                    st.markdown("##### 👔 Supervisor Hrs vs Committed Hrs Pivot")
                    _grp3_sup = (
                        _pdata3
                        .groupby(["Supervisor", "Date (IST)"], sort=True)
                        .agg(
                            _slot_cnt       = ("Scheduled Slot",  "count"),
                            _in3_secs       = ("_in3_secs",       "sum"),
                            _committed_secs = ("_committed_secs", "sum"),
                            _inbound        = ("Inbound Handled",  "sum"),
                            _outbound       = ("Outbound Handled", "sum"),
                        )
                        .reset_index()
                    )
                    _piv3s_dates = sorted(
                        _grp3_sup["Date (IST)"].unique(),
                        key=lambda _d: datetime.strptime(_d, "%d-%b-%Y"),
                    )
                    _piv3s_rows = []
                    for _sup3, _gs3 in _grp3_sup.groupby("Supervisor", sort=True):
                        _rs3 = {"Supervisor": _sup3}
                        _ts_cnt = _ts_in = _ts_comm = _ts_inb = _ts_outb = 0
                        for _ds3 in _piv3s_dates:
                            _dgs3 = _gs3[_gs3["Date (IST)"] == _ds3]
                            _cs3  = int(_dgs3["_slot_cnt"].sum())
                            _is3  = int(_dgs3["_in3_secs"].sum())
                            _co3  = int(_dgs3["_committed_secs"].sum())
                            _rs3[f"{_ds3} | Count of Scheduled Slot"] = _cs3
                            _rs3[f"{_ds3} | Scheduled Hrs"]           = fmt_duration(_co3)
                            _rs3[f"{_ds3} | Sum of In Call Duration"]  = fmt_duration(_is3)
                            _ts_cnt  += _cs3
                            _ts_in   += _is3
                            _ts_comm += _co3
                            _ts_inb  += int(_dgs3["_inbound"].sum())
                            _ts_outb += int(_dgs3["_outbound"].sum())
                        _rs3["Total Count of Scheduled Slot"] = _ts_cnt
                        _rs3["Total Scheduled Hrs"]           = fmt_duration(_ts_comm)
                        _rs3["Total Sum of In Call Duration"]  = fmt_duration(_ts_in)
                        _rs3["Compliance %"] = (
                            _ts_in / _ts_comm
                            if _ts_comm > 0 else 0.0
                        )
                        _rs3["Inbound Handled"]  = _ts_inb
                        _rs3["Outbound Handled"] = _ts_outb
                        _piv3s_rows.append(_rs3)
                    _piv3_sup = pd.DataFrame(_piv3s_rows) if _piv3s_rows else pd.DataFrame()
                    st.dataframe(_piv3_sup, use_container_width=True, hide_index=True)

                    # ── Supervisor AUX Breakdown ────────────────────────────────
                    _sup_aux_pivot = pd.DataFrame()
                    if aux_detail_records:
                        _adt = pd.DataFrame(aux_detail_records)
                        _adt_grp = (
                            _adt.groupby(["Supervisor", "AUX Code"])["_secs"]
                            .sum()
                            .unstack(fill_value=0)
                            .reset_index()
                        )
                        _aux_cols_list = [c for c in _adt_grp.columns if c != "Supervisor"]
                        _adt_grp["Total In Call Duration"] = _adt_grp[_aux_cols_list].sum(axis=1)
                        _sup_aux_pivot = _adt_grp.copy()
                        for _ac in _aux_cols_list + ["Total In Call Duration"]:
                            _sup_aux_pivot[_ac] = _adt_grp[_ac].apply(fmt_duration)
                        st.markdown("##### 📊 Supervisor AUX Breakdown (In Call Duration)")
                        st.dataframe(_sup_aux_pivot, use_container_width=True, hide_index=True)
                else:
                    _piv3_sup = pd.DataFrame()
                    _sup_aux_pivot = pd.DataFrame()

                # ── Detailed table ──────────────────────────────────────────────────
                st.markdown("##### Detailed Compliance Records")
                st.dataframe(report_df, use_container_width=True, hide_index=True)

                # ── Out-of-Slot report ─────────────────────────────────────────
                st.markdown("##### ⏰ Out-of-Slot Login Report")
                if out_records:
                    out_df = pd.DataFrame(out_records)
                    st.caption(
                        f"Logins that occurred **outside** the agent’s scheduled time slot "
                        f"or on Week Off days. "
                        f"({out_df['Agent Name'].nunique()} agent(s), {len(out_df)} record(s))"
                    )
                    oos1, oos2, oos3 = st.columns(3)
                    oos1.metric("Agents with OOS Activity", out_df["Agent Name"].nunique())
                    oos2.metric("Total OOS Records",        len(out_df))
                    oos3.metric("Week Off Logins",
                                int((out_df["Type"] == "Week Off - Logged In").sum()))
                    st.dataframe(out_df, use_container_width=True, hide_index=True)
                else:
                    st.success("✅ No out-of-slot login activity detected.")

                # ── Export ──────────────────────────────────────────────────
                out2 = io.BytesIO()
                with pd.ExcelWriter(out2, engine="openpyxl") as writer:
                    report_df.to_excel(writer, sheet_name="Compliance Report", index=False)
                    _style_ws(writer.sheets["Compliance Report"], pct_col=None)

                    if not sup_summary.empty:
                        _sup_on  = int(sup_summary["✅ On Calls"].sum())
                        _sup_tot = int(sup_summary["Scheduled Slots"].sum())
                        _sup_xl  = pd.concat([
                            sup_summary,
                            pd.DataFrame([{
                                "Supervisor":      "Total",
                                "Agents":          int(sup_summary["Agents"].sum()),
                                "Scheduled Slots": _sup_tot,
                                "✅ On Calls":      _sup_on,
                                "⚠️ No Calls":      int(sup_summary["⚠️ No Calls"].sum()),
                                "❌ Not Logged In": int(sup_summary["❌ Not Logged In"].sum()),
                                "Compliance %":    _sup_on / _sup_tot if _sup_tot > 0 else 0.0,
                            }])
                        ], ignore_index=True)
                        _sup_xl.to_excel(writer, sheet_name="Supervisor Summary", index=False)
                        _style_ws(writer.sheets["Supervisor Summary"], has_total_row=True)
                        _add_bar_chart(writer.sheets["Supervisor Summary"],
                                       "Supervisor", "Supervisor Compliance %",
                                       exclude_last=True)

                    if len(non_wo) > 0:
                        _ag_on  = int(agent_summary["✅ On Calls"].sum())
                        _ag_tot = int(agent_summary["Scheduled Days"].sum())
                        _ag_total_d = {c: "" for c in agent_summary.columns}
                        _ag_total_d.update({
                            "Agent Name":      "Total",
                            "Scheduled Days":  _ag_tot,
                            "✅ On Calls":      _ag_on,
                            "⚠️ No Calls":      int(agent_summary["⚠️ No Calls"].sum()),
                            "❌ Not Logged In": int(agent_summary["❌ Not Logged In"].sum()),
                            "Compliance %":    _ag_on / _ag_tot if _ag_tot > 0 else 0.0,
                        })
                        if "Supervisor" in agent_summary.columns:
                            _ag_total_d["Supervisor"] = "Total"
                        _ag_xl  = pd.concat([
                            agent_summary,
                            pd.DataFrame([_ag_total_d])
                        ], ignore_index=True)
                        _ag_xl.to_excel(writer, sheet_name="Agent Summary", index=False)
                        _style_ws(writer.sheets["Agent Summary"], has_total_row=True)
                        _add_bar_chart(writer.sheets["Agent Summary"],
                                       "Agent Name", "Agent Compliance %",
                                       exclude_last=True)

                    if not _piv_combined.empty:
                        _piv_combined.to_excel(writer, sheet_name="Summary Pivot", index=False)
                        _style_ws(writer.sheets["Summary Pivot"], pct_col=None)

                    if not _piv3.empty:
                        _piv3.to_excel(writer, sheet_name="Hrs vs Committed", index=False)
                        _style_ws(writer.sheets["Hrs vs Committed"], pct_col="Compliance %")

                    if not _piv3_sup.empty:
                        _piv3_sup.to_excel(writer, sheet_name="Sup Hrs vs Committed", index=False)
                        _style_ws(writer.sheets["Sup Hrs vs Committed"], pct_col="Compliance %")

                    if not _sup_aux_pivot.empty:
                        _sup_aux_pivot.to_excel(writer, sheet_name="Sup AUX Breakdown", index=False)
                        _style_ws(writer.sheets["Sup AUX Breakdown"], pct_col=None)

                    if out_records:
                        pd.DataFrame(out_records).to_excel(
                            writer, sheet_name="Out-of-Slot Logins", index=False
                        )
                        _style_ws(writer.sheets["Out-of-Slot Logins"], pct_col=None)

                    sched_raw.to_excel(writer, sheet_name="Raw - Schedule", index=False)
                    _style_ws(writer.sheets["Raw - Schedule"], pct_col=None)

                st.download_button(
                    "📥 Download Compliance Report (Excel)",
                    data=out2.getvalue(),
                    file_name=f"ScheduleCompliance_{date_from}_{date_to}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

# ──────────────────────────────────────────────────────────────────────────────
# OT Tracker
# ──────────────────────────────────────────────────────────────────────────────
st.markdown("---")
st.markdown('<div class="section-header">⏰ OT Tracker</div>', unsafe_allow_html=True)
st.write("")
st.caption(
    "Upload the OT schedule sheet to verify whether agents logged in and handled calls "
    "during their Overtime window. Reuses the login data loaded in the sidebar. "
    "**Eligibility = ✅ agent had In-Call activity during the OT shift.**"
)


def parse_ot_shift(shift_str: str):
    """Parse 24-hour OT shift strings like '19:00-04:00'. Falls back to parse_time_slot."""
    s  = str(shift_str).strip()
    m24 = re.match(r'^(\d{1,2}):(\d{2})\s*[-\u2013]\s*(\d{1,2}):(\d{2})$', s)
    if m24:
        try:
            return time(int(m24.group(1)), int(m24.group(2))), \
                   time(int(m24.group(3)), int(m24.group(4)))
        except ValueError:
            pass
    return parse_time_slot(s)


ot_uploaded = st.file_uploader(
    "📂 Upload OT Schedule (Excel or CSV)",
    type=["xlsx", "xls", "csv"],
    key="ot_uploader",
)
if ot_uploaded is not None:
    _ob = ot_uploaded.read()
    _save_to_disk("ot", _ob, ot_uploaded.name)
    st.session_state["ot_bytes"] = _ob
    st.session_state["ot_name"]  = ot_uploaded.name

if "ot_bytes" not in st.session_state:
    _ob2, _on2 = _load_from_disk("ot")
    if _ob2 is not None:
        st.session_state["ot_bytes"] = _ob2
        st.session_state["ot_name"]  = _on2

if "ot_bytes" in st.session_state:
    _otc1, _otc2 = st.columns([1, 3])
    if _otc1.button("🗑️ Clear OT file", key="clear_ot"):
        _clear_from_disk("ot")
        st.session_state.pop("ot_bytes", None)
        st.session_state.pop("ot_name",  None)
        st.rerun()
    _otc2.caption(f"📄 Using: **{st.session_state['ot_name']}**")

    try:
        ot_raw = load_ot_schedule(st.session_state["ot_bytes"], st.session_state["ot_name"])
    except ValueError as _ot_err:
        st.error(str(_ot_err))
        st.stop()
    st.success(
        f"OT Schedule loaded from **'{OT_SHEET_NAME}'** — "
        f"{len(ot_raw)} rows, {len(ot_raw.columns)} columns"
    )

    with st.expander("Preview OT Schedule", expanded=False):
        st.dataframe(ot_raw, use_container_width=True)

    # ── Column mapping ─────────────────────────────────────────────────────────
    st.markdown("##### Configure OT Schedule Columns")
    ot_cols  = list(ot_raw.columns)
    _ot_idx  = lambda c: (ot_cols.index(c) if c and c in ot_cols else 0)

    _otn_def  = find_col(ot_raw, ["employee name", "agent name", "name"])
    _ots_def  = find_col(ot_raw, ["supervisor"])
    _ote_def  = find_col(ot_raw, ["email id", "email"])
    _otp_def  = find_col(ot_raw, ["process"])

    _login_cols      = list(df.columns)
    _login_email_def = find_col(df, ["agent email", "email id", "email"])
    _li_idx = lambda c: (_login_cols.index(c) if c and c in _login_cols else 0)

    st.caption("🔗 Agents are matched **by Email ID** between the OT schedule and the login data.")
    _ocm1, _ocm2, _ocm3 = st.columns(3)
    with _ocm1:
        ot_name_col = st.selectbox("Employee Name (OT schedule)", ot_cols,
                                   index=_ot_idx(_otn_def))
    with _ocm2:
        ot_email_col = st.selectbox("Email ID — OT schedule ★", ot_cols,
                                    index=_ot_idx(_ote_def))
    with _ocm3:
        ot_login_email_col = st.selectbox("Agent Email — Login data ★", _login_cols,
                                          index=_li_idx(_login_email_def))

    _ocm4, _ocm5 = st.columns(2)
    with _ocm4:
        ot_sup_col = st.selectbox("Supervisor (optional)", ["— skip —"] + ot_cols,
                                  index=(_ot_idx(_ots_def) + 1) if _ots_def else 0)
        if ot_sup_col == "— skip —": ot_sup_col = None
    with _ocm5:
        ot_proc_col = st.selectbox("Process (optional)", ["— skip —"] + ot_cols,
                                   index=(_ot_idx(_otp_def) + 1) if _otp_def else 0)
        if ot_proc_col == "— skip —": ot_proc_col = None

    st.markdown("###### OT Slot Column Sets (up to 3)")
    _ot_triplets = []
    for _n in range(1, 4):
        _wo_def    = find_col(ot_raw, [f"working week off {_n}", f"week off {_n}"])
        _date_def  = find_col(ot_raw, [f"ot working date {_n}", f"ot date {_n}"])
        _shift_def = find_col(ot_raw, [f"ot shift {_n}", f"shift {_n}"])
        _tc1, _tc2, _tc3 = st.columns(3)
        with _tc1:
            _wo = st.selectbox(f"Working Week Off {_n}", ["— skip —"] + ot_cols,
                               index=(_ot_idx(_wo_def) + 1) if _wo_def else 0,
                               key=f"ot_wo_{_n}")
            if _wo == "— skip —": _wo = None
        with _tc2:
            _dt = st.selectbox(f"OT Working Date {_n}", ["— skip —"] + ot_cols,
                               index=(_ot_idx(_date_def) + 1) if _date_def else 0,
                               key=f"ot_date_{_n}")
            if _dt == "— skip —": _dt = None
        with _tc3:
            _sh = st.selectbox(f"OT Shift {_n}", ["— skip —"] + ot_cols,
                               index=(_ot_idx(_shift_def) + 1) if _shift_def else 0,
                               key=f"ot_shift_{_n}")
            if _sh == "— skip —": _sh = None
        if _dt and _sh:
            _ot_triplets.append((_wo, _dt, _sh))

    if not _ot_triplets:
        st.warning("⚠️ Configure at least one OT Date + Shift column pair to continue.")
    else:
        if st.button("▶️ Run OT Compliance Report", type="primary", key="run_ot"):

            # Build email lookup: lower(agent_email) → sub-DataFrame
            _ot_email_lkp: dict = {}
            for _ek, _eg_df in df.groupby(df[ot_login_email_col].str.lower().str.strip()):
                _ot_email_lkp[_ek] = _eg_df

            # Melt wide OT schedule → long format
            ot_long: list = []
            _sno = 0
            for _, ot_row in ot_raw.iterrows():
                emp = str(ot_row[ot_name_col]).strip()
                if not emp or emp.lower() == "nan":
                    continue
                _sup  = str(ot_row[ot_sup_col]).strip()  if ot_sup_col  and pd.notna(ot_row.get(ot_sup_col))  else ""
                _eml  = str(ot_row[ot_email_col]).strip() if pd.notna(ot_row.get(ot_email_col)) else ""
                _prc  = str(ot_row[ot_proc_col]).strip()  if ot_proc_col  and pd.notna(ot_row.get(ot_proc_col))  else ""
                for _wo_col, _date_col, _shift_col in _ot_triplets:
                    _rd = ot_row.get(_date_col)
                    _rs = ot_row.get(_shift_col)
                    _rw = str(ot_row.get(_wo_col, "")).strip() if _wo_col else ""
                    if pd.isna(_rd) or str(_rd).strip().lower() in ("", "nan"):
                        continue
                    if pd.isna(_rs) or str(_rs).strip().lower() in ("", "nan"):
                        continue
                    _sno += 1
                    ot_long.append({
                        "S.No":          _sno,
                        "Employee Name": emp,
                        "Email ID":      _eml,
                        "Supervisor":    _sup,
                        "Process":       _prc,
                        "Week Off Day":  _rw,
                        "_raw_date":     _rd,
                        "_raw_shift":    str(_rs).strip(),
                    })

            if not ot_long:
                st.warning("No OT slots found — check column mapping.")
            else:
                ot_records   = []
                _ot_unmatched: set = set()
                _ot_prog = st.progress(0.0, text="Processing OT slots…")

                for _i, _slot in enumerate(ot_long):
                    _ot_prog.progress(min((_i + 1) / len(ot_long), 1.0))
                    _emp     = _slot["Employee Name"]
                    _eml_key = _slot["Email ID"].lower().strip()

                    def _ot_base():
                        return {k: v for k, v in _slot.items()
                                if k not in ("_raw_date", "_raw_shift")}

                    # Parse OT date
                    try:
                        _rd2 = _slot["_raw_date"]
                        _ot_date = (pd.Timestamp(_rd2).date()
                                    if isinstance(_rd2, (datetime, pd.Timestamp))
                                    else pd.to_datetime(str(_rd2), dayfirst=True).date())
                    except Exception:
                        ot_records.append({**_ot_base(),
                                           "OT Date":              str(_slot["_raw_date"]),
                                           "OT Shift":             _slot["_raw_shift"],
                                           "Status":               "⚠️ Date Parse Error",
                                           "In Call Duration":     "-",
                                           "Total Active in Slot": "-",
                                           "AUX Breakdown":        ""})
                        continue

                    _ot_date_str = _ot_date.strftime("%d-%b-%Y")

                    # Parse shift
                    _parsed_shift = parse_ot_shift(_slot["_raw_shift"])
                    if _parsed_shift is None:
                        ot_records.append({**_ot_base(),
                                           "OT Date":              _ot_date_str,
                                           "OT Shift":             _slot["_raw_shift"],
                                           "Status":               "⚠️ Shift Parse Error",
                                           "In Call Duration":     "-",
                                           "Total Active in Slot": "-",
                                           "AUX Breakdown":        ""})
                        continue

                    _st, _et = _parsed_shift
                    _qs = datetime.combine(_ot_date, _st)
                    _qe = datetime.combine(_ot_date, _et)
                    if _qe <= _qs:
                        _qe += timedelta(days=1)
                    _next_date = _ot_date + timedelta(days=1)

                    # Find agent rows in login data by email
                    if not _eml_key:
                        _ot_unmatched.add(f"{_emp} (no email)")
                        ot_records.append({**_ot_base(),
                                           "OT Date":              _ot_date_str,
                                           "OT Shift":             _slot["_raw_shift"],
                                           "Status":               "⚠️ Email Missing in OT Schedule",
                                           "In Call Duration":     "-",
                                           "Total Active in Slot": "-",
                                           "AUX Breakdown":        ""})
                        continue
                    _ag_rows = _ot_email_lkp.get(_eml_key)
                    if _ag_rows is None:
                        _ot_unmatched.add(f"{_emp} <{_slot['Email ID']}>")
                        ot_records.append({**_ot_base(),
                                           "OT Date":              _ot_date_str,
                                           "OT Shift":             _slot["_raw_shift"],
                                           "Status":               "❌ Agent Not Found in Login Data",
                                           "In Call Duration":     "-",
                                           "Total Active in Slot": "-",
                                           "AUX Breakdown":        ""})
                        continue

                    # Overlap within OT window (include next date for midnight-crossing)
                    _day_rows2 = _ag_rows[
                        _ag_rows[login_col].dt.date.isin([_ot_date, _next_date])
                    ]
                    _ov2 = _day_rows2[
                        (_day_rows2[login_col] < _qe) & (_day_rows2[logout_col] > _qs)
                    ].copy()

                    _b2 = {**_ot_base(), "OT Date": _ot_date_str, "OT Shift": _slot["_raw_shift"]}

                    if _ov2.empty:
                        ot_records.append({**_b2,
                                           "Status":               "❌ Not Logged In",
                                           "In Call Duration":     "00:00:00",
                                           "Total Active in Slot": "00:00:00",
                                           "AUX Breakdown":        "No activity"})
                        continue

                    # Precise logout via Duration column
                    if dur_default and dur_default in _ov2.columns:
                        _ds2 = _ov2[dur_default].apply(_parse_dur_secs)
                        _hd2 = _ds2.notna() & (_ds2 > 0)
                        _pe2 = _ov2[logout_col].copy()
                        _pe2[_hd2] = (
                            _ov2.loc[_hd2, login_col]
                            + _ds2[_hd2].apply(lambda s: timedelta(seconds=s))
                        )
                    else:
                        _pe2 = _ov2[logout_col]

                    _ov2["_es"] = _ov2[login_col].clip(lower=_qs, upper=_qe)
                    _ov2["_ee"] = _pe2.clip(lower=_qs, upper=_qe)
                    _ov2 = _ov2[_ov2["_ee"] > _ov2["_es"]].copy()

                    if _ov2.empty:
                        ot_records.append({**_b2,
                                           "Status":               "❌ Not Logged In",
                                           "In Call Duration":     "00:00:00",
                                           "Total Active in Slot": "00:00:00",
                                           "AUX Breakdown":        "No activity"})
                        continue

                    _all_iv2   = list(zip(_ov2["_es"], _ov2["_ee"]))
                    _tot2      = _union_secs(_all_iv2)
                    # In Call Duration = actual call time only (skill-name records = "In Call - Working")
                    _call_mask = _ov2["_aux_label"].str.lower().str.strip().isin(
                        {"in call - working", "in call- working"}
                    )
                    _wk_iv2    = list(zip(_ov2.loc[_call_mask, "_es"], _ov2.loc[_call_mask, "_ee"]))
                    _inc2      = _union_secs(_wk_iv2)

                    _tr2       = list(zip(_ov2["_aux_label"], _ov2["_es"], _ov2["_ee"]))
                    _ag2       = pd.Series(_allocate_aux_secs(_tr2)).sort_values(ascending=False)
                    _ax2       = "; ".join(f"{k}: {fmt_duration(v)}"
                                           for k, v in _ag2.items() if v > 0)

                    _st2 = ("✅ Eligible - On Calls"
                            if _inc2 > 0 else "⚠️ Logged In - No Calls")
                    ot_records.append({**_b2,
                                       "Status":               _st2,
                                       "In Call Duration":     fmt_duration(_inc2),
                                       "Total Active in Slot": fmt_duration(_tot2),
                                       "AUX Breakdown":        _ax2})

                _ot_prog.empty()

                if ot_records:
                    ot_df = pd.DataFrame(ot_records)

                    _ot_elig = int(ot_df["Status"].str.startswith("✅").sum())
                    _ot_nc   = int(ot_df["Status"].str.startswith("⚠️ Logged").sum())
                    _ot_ni   = int(ot_df["Status"].str.startswith("❌ Not L").sum())
                    _ot_err  = int(ot_df["Status"].str.contains(
                        "Error|Not Found", na=False).sum())

                    _om1, _om2, _om3, _om4, _om5 = st.columns(5)
                    _om1.metric("Total OT Slots",         len(ot_df))
                    _om2.metric("✅ Eligible (On Calls)",  _ot_elig)
                    _om3.metric("⚠️ Logged - No Calls",    _ot_nc)
                    _om4.metric("❌ Not Logged In",         _ot_ni)
                    _om5.metric("⚠️ Errors / Not Found",   _ot_err)

                    if _ot_unmatched:
                        st.warning(
                            f"⚠️ {len(_ot_unmatched)} OT entry/entries not matched in login data — "
                            "check email addresses: "
                            + ", ".join(sorted(_ot_unmatched))
                        )

                    st.dataframe(ot_df, use_container_width=True, hide_index=True)

                    _ot_out = io.BytesIO()
                    with pd.ExcelWriter(_ot_out, engine="openpyxl") as _ot_writer:
                        ot_df.to_excel(_ot_writer, sheet_name="OT Compliance", index=False)
                        _style_ws(_ot_writer.sheets["OT Compliance"], pct_col=None)
                        ot_raw.to_excel(_ot_writer, sheet_name="Raw - OT Schedule", index=False)
                        _style_ws(_ot_writer.sheets["Raw - OT Schedule"], pct_col=None)

                    st.download_button(
                        "📥 Download OT Compliance Report (Excel)",
                        data=_ot_out.getvalue(),
                        file_name=f"OT_Compliance_{date.today().strftime('%d-%b-%Y')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
else:
    st.info("📤 Upload an OT schedule file above to get started.")
